import os
import json
import time
import base64
import requests
import re
import threading
import pandas as pd
import sqlalchemy
from sqlalchemy import text
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from flask import request, jsonify
from .storage import storage
from .utils import log, resolve_variables, get_ai

def export_to_excel(data, node_id, execution_id):
    """Export query result to Excel with auto-fit columns and highlighted headers."""
    try:
        if not data or not isinstance(data, list):
            return None
            
        df = pd.DataFrame(data)
        file_path = f"/tmp/query_result_{execution_id}_{node_id}.xlsx"
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Query Results')
            workbook = writer.book
            worksheet = writer.sheets['Query Results']
            
            # Highlight headers in yellow
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            header_font = Font(bold=True)
            
            for cell in worksheet[1]:
                cell.fill = yellow_fill
                cell.font = header_font
            
            # Auto-fit column width
            for i, col in enumerate(df.columns):
                column_len = max(df[col].astype(str).str.len().max(), len(col)) + 2
                worksheet.column_dimensions[get_column_letter(i + 1)].width = column_len
                
        return file_path
    except Exception as e:
        log(f"Excel export failed: {e}")
        return None

def get_dag_state(dag_id, base_url, auth_headers):
    try:
        response = requests.get(
            f"{base_url}/api/v1/dags/{dag_id}/dagRuns",
            params={'order_by': '-execution_date', 'limit': 1},
            headers=auth_headers
        )
        response.raise_for_status()
        dag_runs = response.json().get('dag_runs', [])
        if dag_runs:
            return dag_runs[0].get('state', 'unknown')
        return 'no_runs'
    except Exception as e:
        log(f"Error checking DAG state for {dag_id}: {e}")
        return 'unknown'

def wait_for_dags_to_complete(dag_infos, logs, execution_id, storage):
    max_wait_time = 3600
    poll_interval = 10
    elapsed = 0
    
    while elapsed < max_wait_time:
        all_complete = True
        for dag_info in dag_infos:
            dag_id = dag_info['dag_id']
            base_url = dag_info['base_url']
            auth_headers = dag_info['auth_headers']
            
            if not base_url:
                continue
            
            state = get_dag_state(dag_id, base_url, auth_headers)
            running_states = ['running', 'queued', 'scheduled', 'up_for_retry', 'up_for_reschedule', 'restarting', 'deferred']
            
            if state.lower() in running_states:
                all_complete = False
                logs.append({
                    'timestamp': datetime.now().isoformat(),
                    'level': 'INFO',
                    'message': f"DAG {dag_id} is currently {state}. Waiting for it to reach a terminal state..."
                })
                storage.update_execution(execution_id, 'waiting', logs)
                break
        
        if all_complete:
            return True
        
        time.sleep(poll_interval)
        elapsed += poll_interval
    
    logs.append({
        'timestamp': datetime.now().isoformat(),
        'level': 'ERROR',
        'message': f"Timeout waiting for DAGs to reach a terminal state after {max_wait_time} seconds"
    })
    return False

def collect_dag_infos_from_workflow(nodes, storage):
    dag_infos = []
    for node in nodes:
        node_data = node.get('data', {})
        node_type = node_data.get('type')
        
        if node_type in ['airflow_trigger', 'airflow_log_check']:
            config = node_data.get('config', {})
            dag_id = config.get('dagId', '')
            credential_id = config.get('credentialId')
            
            base_url = ""
            auth_headers = {}
            
            if credential_id:
                cred = storage.get_credential(int(credential_id))
                if cred and cred.get('type') == 'airflow':
                    cred_data = cred.get('data', {})
                    base_url = cred_data.get('baseUrl', '')
                    auth = base64.b64encode(f"{cred_data.get('username')}:{cred_data.get('password')}".encode()).decode()
                    auth_headers = {'Authorization': f'Basic {auth}'}
            
            if dag_id and base_url:
                dag_infos.append({
                    'dag_id': dag_id,
                    'base_url': base_url,
                    'auth_headers': auth_headers
                })
    return dag_infos

def execute_workflow_async(execution_id, workflow_id):
    workflow = storage.get_workflow(workflow_id)
    if not workflow:
        return
    
    logs = []
    results = {}
    nodes = workflow.get('nodes', [])
    edges = workflow.get('edges', [])
    
    logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': 'Checking if any involved DAGs are currently running...'})
    storage.update_execution(execution_id, 'checking', logs)
    
    dag_infos = collect_dag_infos_from_workflow(nodes, storage)
    
    if dag_infos:
        if not wait_for_dags_to_complete(dag_infos, logs, execution_id, storage):
            logs.append({'timestamp': datetime.now().isoformat(), 'level': 'ERROR', 'message': 'Workflow aborted: DAGs did not complete in time'})
            storage.update_execution(execution_id, 'failed', logs, results)
            return
    
    logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': 'Starting workflow execution...'})
    storage.update_execution(execution_id, 'running', logs)
    
    execution_context = {}
    
    def find_next_nodes(current_node_id, handle=None):
        return [
            node for edge in edges
            if edge.get('source') == current_node_id and (not handle or edge.get('sourceHandle') == handle)
            for node in nodes if node.get('id') == edge.get('target')
        ]
    
    current_nodes = [n for n in nodes if not any(e.get('target') == n.get('id') for e in edges)]
    visited = set()
    assertion_failed = False
    
    while current_nodes and not assertion_failed:
        next_batch = []
        for node in current_nodes:
            node_id = node.get('id')
            if node_id in visited: continue
            visited.add(node_id)
            
            time.sleep(1)
            node_data = node.get('data', {})
            node_type = node_data.get('type')
            config = node_data.get('config', {})
            
            logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': f"Executing node {node_data.get('label')} ({node_type})..."})
            results[node_id] = {'status': 'running'}
            storage.update_execution(execution_id, 'running', logs, results)
            
            output_handle = 'output'
            try:
                if node_type == 'airflow_trigger':
                    dag_id = resolve_variables(config.get('dagId', ''), execution_context)
                    conf = config.get('conf', {})
                    credential_id = config.get('credentialId')
                    wait_for_completion = config.get('waitForCompletion', True)
                    
                    auth_headers = {}
                    base_url = ""
                    username = ""
                    password = ""
                    
                    if credential_id:
                        cred = storage.get_credential(int(credential_id))
                        if cred and cred.get('type') == 'airflow':
                            cred_data = cred.get('data', {})
                            base_url = cred_data.get('baseUrl', '')
                            username = cred_data.get('username', '')
                            password = cred_data.get('password', '')
                            auth = base64.b64encode(f"{username}:{password}".encode()).decode()
                            auth_headers = {'Authorization': f'Basic {auth}'}
                    
                    dag_run_id = f"run_{int(time.time() * 1000)}"
                    if base_url:
                        response = requests.post(f"{base_url}/api/v1/dags/{dag_id}/dagRuns", json={'conf': conf}, headers=auth_headers)
                        response.raise_for_status()
                        dag_run_id = response.json().get('dag_run_id', dag_run_id)
                        
                        if wait_for_completion:
                            logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': f"Waiting for DAG {dag_id} (run: {dag_run_id}) to complete..."})
                            terminal_states = ['success', 'failed']
                            max_wait_time = 3600  # 1 hour timeout
                            poll_interval = 10
                            elapsed_wait = 0
                            
                            while elapsed_wait < max_wait_time:
                                try:
                                    run_response = requests.get(f"{base_url}/api/v1/dags/{dag_id}/dagRuns/{dag_run_id}", headers=auth_headers)
                                    run_response.raise_for_status()
                                    current_state = run_response.json().get('state', 'unknown')
                                    
                                    if current_state.lower() in terminal_states:
                                        logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': f"DAG {dag_id} finished with state: {current_state}"})
                                        if current_state.lower() == 'failed':
                                            raise Exception(f"DAG {dag_id} failed")
                                        break
                                        
                                    # Log progress every minute
                                    if elapsed_wait % 60 == 0:
                                        logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': f"DAG {dag_id} is still {current_state}..."})
                                        storage.update_execution(execution_id, 'running', logs, results)
                                        
                                except Exception as poll_error:
                                    log(f"Error polling DAG status: {poll_error}")
                                    # Don't fail immediately on network blips
                                
                                time.sleep(poll_interval)
                                elapsed_wait += poll_interval
                            else:
                                raise Exception(f"Timeout waiting for DAG {dag_id} to complete after {max_wait_time} seconds")
                    
                    execution_context['dagRunId'] = dag_run_id
                    execution_context['dagId'] = dag_id
                    
                    # Store result in node-specific context as well
                    execution_context[node_id] = {
                        'dagId': dag_id,
                        'dagRunId': dag_run_id,
                        'status': 'success'
                    }
                    results[node_id] = {'status': 'success', 'dagId': dag_id, 'dagRunId': dag_run_id}
                
                elif node_type == 'airflow_log_check':
                    node_dag_id = resolve_variables(config.get('dagId', execution_context.get('dagId', '')), execution_context)
                    task_name = resolve_variables(config.get('taskName', ''), execution_context)
                    log_assertions = config.get('logAssertions', []) # Expecting a list of strings
                    if not log_assertions and config.get('logAssertion'):
                        log_assertions = [config.get('logAssertion')]
                    
                    run_id = execution_context.get('dagRunId', '')
                    credential_id = config.get('credentialId') or next((n.get('data', {}).get('config', {}).get('credentialId') for n in nodes if n.get('data', {}).get('type') == 'airflow_trigger'), None)
                    
                    if not run_id or not node_dag_id or not task_name:
                        raise Exception("Missing DAG ID, Task Name, or Run ID for log check")

                    auth_headers = {}
                    base_url = ""
                    if credential_id:
                        cred = storage.get_credential(int(credential_id))
                        if cred and cred.get('type') == 'airflow':
                            cred_data = cred.get('data', {})
                            base_url = cred_data.get('baseUrl', '')
                            auth = base64.b64encode(f"{cred_data.get('username')}:{cred_data.get('password')}".encode()).decode()
                            auth_headers = {'Authorization': f'Basic {auth}'}

                    if base_url:
                        # Get task logs
                        log_response = requests.get(f"{base_url}/api/v1/dags/{node_dag_id}/dagRuns/{run_id}/taskInstances/{task_name}/logs/1", headers=auth_headers)
                        log_response.raise_for_status()
                        logs_text = log_response.text
                        
                        failed_assertions = []
                        for assertion in log_assertions:
                            resolved_assertion = resolve_variables(assertion, execution_context)
                            if resolved_assertion not in logs_text:
                                failed_assertions.append(resolved_assertion)
                        
                        if failed_assertions:
                            error_msg = f"Assertions failed: {', '.join(failed_assertions)}"
                            logs.append({'timestamp': datetime.now().isoformat(), 'level': 'ERROR', 'message': error_msg})
                            results[node_id] = {'status': 'failure', 'error': error_msg}
                            assertion_failed = True
                            break
                        else:
                            logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': f"All {len(log_assertions)} log assertions passed for task {task_name}"})
                            execution_context[node_id] = {'status': 'success'}
                            results[node_id] = {'status': 'success'}
                    else:
                        raise Exception("Airflow credential not found for log check")
                
                elif node_type == 'parallel_dags':
                    dag_configs = config.get('dags', [])
                    threads = []
                    parallel_results = {}
                    
                    def run_single_dag(dag_conf, idx):
                        # Simulating trigger logic for each dag in the list
                        # In a real scenario, this would call airflow_trigger logic
                        # For now, we'll log it
                        dag_id = resolve_variables(dag_conf.get('dagId', ''), execution_context)
                        logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': f"Parallel trigger: {dag_id}"})
                        parallel_results[f"dag_{idx}"] = "triggered"

                    for i, dag_conf in enumerate(dag_configs):
                        t = threading.Thread(target=run_single_dag, args=(dag_conf, i))
                        t.start()
                        threads.append(t)
                    
                    for t in threads:
                        t.join()
                        
                    results[node_id] = {'status': 'success', 'parallel_results': parallel_results}

                elif node_type == 'sql_query':
                    query = resolve_variables(config.get('query', ''), execution_context)
                    credential_id = config.get('credentialId')
                    logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': f"Running SQL: {query}"})
                    
                    query_results = []
                    try:
                        if credential_id:
                            cred = storage.get_credential(int(credential_id))
                            if cred:
                                cred_type = cred.get('type')
                                cred_data = cred.get('data', {})
                                
                                if cred_type == 'mssql':
                                    conn_str = f"mssql+pymssql://{cred_data.get('username')}:{cred_data.get('password')}@{cred_data.get('host')}:{cred_data.get('port', 1433)}/{cred_data.get('database')}"
                                    engine = sqlalchemy.create_engine(conn_str)
                                    with engine.connect() as conn:
                                        result = conn.execute(text(query))
                                        query_results = [dict(row._mapping) for row in result]
                                elif cred_type == 'postgres':
                                    conn_str = f"postgresql://{cred_data.get('username')}:{cred_data.get('password')}@{cred_data.get('host')}:{cred_data.get('port', 5432)}/{cred_data.get('database')}"
                                    engine = sqlalchemy.create_engine(conn_str)
                                    with engine.connect() as conn:
                                        result = conn.execute(text(query))
                                        query_results = [dict(row._mapping) for row in result]
                                else:
                                    raise Exception(f"Unsupported SQL credential type: {cred_type}")
                        else:
                            # Use internal database engine
                            from .models import engine as internal_engine
                            with internal_engine.connect() as conn:
                                result = conn.execute(text(query))
                                query_results = [dict(row._mapping) for row in result]
                    except Exception as e:
                        log(f"SQL Execution failed: {e}")
                        raise Exception(f"SQL Error: {str(e)}")

                    excel_path = export_to_excel(query_results, node_id, execution_id)
                    if excel_path:
                        logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': f"Query results exported to Excel: {excel_path}"})
                    
                    record_count = len(query_results)
                    execution_context['queryResult'] = {'record_count': record_count}
                    execution_context[node_id] = {'count': record_count, 'excel_path': excel_path, 'results': query_results}
                    results[node_id] = {'status': 'success', 'count': record_count, 'excel_path': excel_path}
                
                elif node_type == 'api_request':
                    url = resolve_variables(config.get('url', ''), execution_context)
                    method = config.get('method', 'GET').upper()
                    headers = config.get('headers', {})
                    body = resolve_variables(config.get('body', ''), execution_context)
                    
                    logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': f"Sending {method} request to {url}"})
                    response = requests.request(method, url, headers=headers, data=body)
                    response.raise_for_status()
                    
                    try:
                        res_data = response.json()
                    except:
                        res_data = response.text
                        
                    execution_context[node_id] = {'response': res_data}
                    results[node_id] = {'status': 'success', 'data': res_data}

                elif node_type == 'python_script':
                    script_code = config.get('code', '')
                    logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO', 'message': "Executing Python script..."})
                    
                    local_scope = {'context': execution_context, 'result': None, 'requests': requests, 'json': json, 'pd': pd}
                    exec(script_code, {}, local_scope)
                    
                    script_result = local_scope.get('result')
                    execution_context[node_id] = {'result': script_result}
                    results[node_id] = {'status': 'success', 'result': script_result}
                
                elif node_type == 's3_operation':
                    bucket = resolve_variables(config.get('bucket', ''), execution_context)
                    operation = config.get('operation', 'list')
                    key = resolve_variables(config.get('key', ''), execution_context)
                    credential_id = config.get('credentialId')
                    
                    if not credential_id:
                        raise Exception("S3 credentials required")
                    
                    cred = storage.get_credential(int(credential_id))
                    if not cred or cred.get('type') != 's3':
                        raise Exception("Invalid S3 credential")
                    
                    cred_data = cred.get('data', {})
                    import boto3
                    s3 = boto3.client(
                        's3',
                        aws_access_key_id=cred_data.get('accessKey'),
                        aws_secret_access_key=cred_data.get('secretKey'),
                        region_name=cred_data.get('region', 'us-east-1')
                    )
                    
                    if operation == 'list':
                        response = s3.list_objects_v2(Bucket=bucket, Prefix=config.get('prefix', ''))
                        contents = [obj['Key'] for obj in response.get('Contents', [])]
                        execution_context[node_id] = {'files': contents}
                        results[node_id] = {'status': 'success', 'files': contents}
                    elif operation == 'upload':
                        content = resolve_variables(config.get('content', ''), execution_context)
                        s3.put_object(Bucket=bucket, Key=key, Body=content)
                        results[node_id] = {'status': 'success'}
                    elif operation == 'delete':
                        s3.delete_object(Bucket=bucket, Key=key)
                        results[node_id] = {'status': 'success'}
                    
                elif node_type == 'sftp_operation':
                    host = resolve_variables(config.get('host', ''), execution_context)
                    port = int(config.get('port', 22))
                    operation = config.get('operation', 'list') # list, upload, download, delete
                    remote_path = resolve_variables(config.get('remotePath', ''), execution_context)
                    credential_id = config.get('credentialId')
                    
                    if not credential_id:
                        raise Exception("SFTP credentials required")
                    
                    cred = storage.get_credential(int(credential_id))
                    if not cred or cred.get('type') != 'sftp':
                        raise Exception("Invalid SFTP credential")
                    
                    cred_data = cred.get('data', {})
                    import paramiko
                    transport = paramiko.Transport((host, port))
                    transport.connect(username=cred_data.get('username'), password=cred_data.get('password'))
                    sftp = paramiko.SFTPClient.from_transport(transport)
                    
                    try:
                        if operation == 'list':
                            files = sftp.listdir(remote_path or '.')
                            execution_context[node_id] = {'files': files}
                            results[node_id] = {'status': 'success', 'files': files}
                        elif operation == 'upload':
                            content = resolve_variables(config.get('content', ''), execution_context)
                            import io
                            file_obj = io.BytesIO(content.encode() if isinstance(content, str) else content)
                            sftp.putfo(file_obj, remote_path)
                            results[node_id] = {'status': 'success'}
                        elif operation == 'download':
                            import io
                            file_obj = io.BytesIO()
                            sftp.getfo(remote_path, file_obj)
                            file_obj.seek(0)
                            downloaded_content = file_obj.read().decode()
                            execution_context[node_id] = {'content': downloaded_content}
                            results[node_id] = {'status': 'success', 'content': downloaded_content}
                        elif operation == 'delete':
                            sftp.remove(remote_path)
                            results[node_id] = {'status': 'success'}
                    finally:
                        sftp.close()
                        transport.close()
                    
                next_batch.extend(find_next_nodes(node_id, output_handle if node_type == 'condition' else None))
            except Exception as e:
                logs.append({'timestamp': datetime.now().isoformat(), 'level': 'ERROR', 'message': f"Error: {e}"})
                results[node_id] = {'status': 'failure', 'error': str(e)}
                assertion_failed = True
                break
            
            storage.update_execution(execution_id, 'running', logs, results)
        current_nodes = next_batch
    
    final_status = 'failed' if assertion_failed else 'completed'
    logs.append({'timestamp': datetime.now().isoformat(), 'level': 'INFO' if not assertion_failed else 'ERROR', 'message': f'Workflow {final_status}.'})
    storage.update_execution(execution_id, final_status, logs, results)

def generate_python_code(workflow):
    nodes = workflow.get('nodes', [])
    edges = workflow.get('edges', [])
    
    code = [
        "import requests",
        "import base64",
        "import json",
        "import time",
        "import boto3",
        "import paramiko",
        "import io",
        "from datetime import datetime",
        "",
        "execution_context = {}",
        "",
        "def resolve_variables(text, context):",
        "    if not isinstance(text, str): return text",
        "    for key, value in context.items():",
        "        text = text.replace('{{' + key + '}}', str(value))",
        "    return text",
        ""
    ]
    
    # Simple topological sort for sequence
    # This is a basic implementation; for complex graphs, a real topo sort is needed
    for node in nodes:
        node_id = node.get('id')
        node_data = node.get('data', {})
        node_type = node_data.get('type')
        config = node_data.get('config', {})
        
        code.append(f"# Node: {node_data.get('label')} ({node_type})")
        
        if node_type == 'airflow_trigger':
            code.append(f"print('Triggering DAG: {config.get('dagId')}')")
            # Code to trigger DAG would go here using requests
        elif node_type == 'sql_query':
            code.append(f"print('Running SQL: {config.get('query')}')")
        elif node_type == 's3_operation':
            code.append(f"print('S3 Operation: {config.get('operation')} on {config.get('bucket')}')")
        elif node_type == 'sftp_operation':
            code.append(f"print('SFTP Operation: {config.get('operation')} on {config.get('host')}')")
        
        code.append("")
        
    return "\n".join(code)

def generate_pytest_suite(workflow_ids):
    import json
    
    code = [
        "import pytest",
        "import requests",
        "import base64",
        "import json",
        "import time",
        "import boto3",
        "import paramiko",
        "from datetime import datetime",
        "",
        "class AutomationSuite:",
        "    def __init__(self):",
        "        self.context = {}",
        "",
        "    def resolve(self, text):",
        "        if not isinstance(text, str): return text",
        "        for key, value in self.context.items():",
        "            text = text.replace('{{' + key + '}}', str(value))",
        "        return text",
        "",
        "    def run_airflow_trigger(self, config):",
        "        print(f\"Triggering DAG: {config.get('dagId')}\")",
        "        # Implementation details...",
        "        return \"run_123\"",
        "",
        "    def run_sql_query(self, config):",
        "        print(f\"Running SQL: {config.get('query')}\")",
        "        return {\"count\": 101}",
        ""
    ]
    
    for wid in workflow_ids:
        workflow = storage.get_workflow(wid)
        if not workflow: continue
        
        test_name = workflow['name'].lower().replace(' ', '_')
        code.append(f"def test_{test_name}():")
        code.append("    suite = AutomationSuite()")
        
        nodes = json.loads(workflow.get('nodes', '[]'))
        for node in nodes:
            node_data = node.get('data', {})
            node_type = node_data.get('type')
            config = node_data.get('config', {})
            
            if node_type == 'airflow_trigger':
                code.append(f"    suite.context['{node['id']}'] = suite.run_airflow_trigger({config})")
            elif node_type == 'sql_query':
                code.append(f"    result = suite.run_sql_query({config})")
                code.append(f"    assert result['count'] > 0")
        code.append("")
        
    return "\n".join(code)

def register_workflow_routes(app):
    @app.post('/api/automation/export-suite')
    def export_automation_suite():
        data = request.get_json()
        ids = data.get('workflowIds', [])
        code = generate_pytest_suite(ids)
        return jsonify({'code': code})

    @app.post('/api/git/sync')
    def git_sync():
        import git
        data = request.get_json()
        action = data.get('action') # 'push' or 'pull'
        repo_path = os.getcwd()
        
        try:
            repo = git.Repo(repo_path)
            if action == 'push':
                repo.git.add(A=True)
                repo.index.commit("Automation Suite Update")
                origin = repo.remote(name='origin')
                origin.push()
            elif action == 'pull':
                origin = repo.remote(name='origin')
                origin.pull()
            return jsonify({'status': 'success'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500

    @app.get('/api/workflows/<int:id>/export')
    def export_workflow(id):
        workflow = storage.get_workflow(id)
        if not workflow: return jsonify({'message': 'Workflow not found'}), 404
        python_code = generate_python_code(workflow)
        return jsonify({'code': python_code})

    @app.get('/api/workflows')
    def list_workflows():
        return jsonify(storage.get_workflows())

    @app.get('/api/workflows/<int:id>')
    def get_workflow(id):
        workflow = storage.get_workflow(id)
        if not workflow: return jsonify({'message': 'Workflow not found'}), 404
        return jsonify(workflow)

    @app.post('/api/workflows')
    def create_workflow():
        return jsonify(storage.create_workflow(request.get_json())), 201

    @app.put('/api/workflows/<int:id>')
    def update_workflow(id):
        workflow = storage.update_workflow(id, request.get_json())
        if not workflow: return jsonify({'message': 'Workflow not found'}), 404
        return jsonify(workflow)

    @app.delete('/api/workflows/<int:id>')
    def delete_workflow(id):
        storage.delete_workflow(id)
        return '', 204

    @app.post('/api/workflows/generate')
    def generate_workflow():
        data = request.get_json()
        prompt = data.get('prompt', '')
        workflow_id = data.get('workflowId')
        
        max_retries = 2
        for attempt in range(max_retries):
            try:
                openai = get_ai()
                response = openai.chat.completions.create(
                    model="gpt-4o",
                    messages=[
                        {
                            "role": "system",
                            "content": """You are a workflow generator for Apache Airflow 2.7.3 and SQL Server.
CRITICAL: Return ONLY a JSON object with 'nodes' and 'edges' compatible with React Flow.
ONE OPERATION PER NODE. Nodes can pass values using double curly braces (e.g., {{dagRunId}}, {{queryResult}}).

Available node types:
- 'airflow_trigger': { dagId: string, conf?: object }. Output: dagRunId.
- 'airflow_log_check': { dagId: string, taskName?: string, logAssertion: string }.
- 'sql_query': { query: string, credentialId: number }. Output: queryResult.
- 'condition': { threshold: number, variable: string, operator: string }. 
  CRITICAL: Use sourceHandle "success" for true and "failure" for false in outgoing edges.
- 'api_request': { url: string, method: string, headers: object, body: string }.
- 'python_script': { code: string }.

Example response format:
{
  "nodes": [
    { "id": "1", "type": "airflow_trigger", "data": { "label": "Trigger", "type": "airflow_trigger", "config": { "dagId": "test" } }, "position": { "x": 0, "y": 0 } }
  ],
  "edges": []
}"""
                        },
                        {"role": "user", "content": prompt}
                    ],
                    response_format={"type": "json_object"},
                    max_completion_tokens=2048
                )
                
                raw_content = response.choices[0].message.content or "{}"
                log(f"AI Raw Response: {raw_content}")
                content = json.loads(raw_content)
                
                if 'nodes' not in content or not isinstance(content['nodes'], list):
                    raise ValueError("Invalid response: 'nodes' array is missing")
                
                if workflow_id:
                    storage.update_workflow(int(workflow_id), {'lastPrompt': prompt})
                
                return jsonify(content)
            except Exception as e:
                log(f"AI Generation attempt {attempt + 1} failed: {e}")
                if attempt == max_retries - 1:
                    return jsonify({
                        'message': 'Failed to generate workflow after multiple attempts',
                        'error': str(e)
                    }), 500
                time.sleep(0.5 * (attempt + 1))

    @app.post('/api/workflows/<int:id>/execute')
    def execute_workflow(id):
        execution = storage.create_execution(id)
        thread = threading.Thread(target=execute_workflow_async, args=(execution['id'], id))
        thread.start()
        return jsonify(execution), 201
